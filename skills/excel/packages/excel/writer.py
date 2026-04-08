#!/usr/bin/env python3
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
TITLE_FONT = Font(bold=True, size=14)
TITLE_ALIGNMENT = Alignment(horizontal="center", vertical="center")
DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color="2F5496"),
    right=Side(style="thin", color="2F5496"),
    top=Side(style="thin", color="2F5496"),
    bottom=Side(style="medium", color="2F5496"),
)
ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)


def estimate_width(text: str) -> float:
    if not text:
        return 0
    width = 0.0
    for ch in str(text):
        if ord(ch) > 127:
            width += 2.0
        else:
            width += 1.0
    return width


def convert_value(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        try:
            if "." in value:
                return float(value)
            return int(value)
        except (ValueError, TypeError):
            return value
    return value


def get_num_cols(sheet_config: dict) -> int:
    headers = sheet_config.get("headers", [])
    data = sheet_config.get("data", [])
    if headers:
        return len(headers)
    if data:
        return max(len(row) for row in data)
    return 0


def apply_cell_style(cell, fill=None, font=None, alignment=None, border=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def create_sheet(wb: Workbook, sheet_config: dict):
    name = sheet_config.get("name", "Sheet1")
    ws = wb.create_sheet(title=name)

    title = sheet_config.get("title")
    headers = sheet_config.get("headers", [])
    data = sheet_config.get("data", [])
    column_widths = sheet_config.get("column_widths")
    merge_cells = sheet_config.get("merge_cells", [])

    num_cols = get_num_cols(sheet_config)
    current_row = 1

    if title:
        cell = ws.cell(row=1, column=1, value=title)
        apply_cell_style(cell, font=TITLE_FONT, alignment=TITLE_ALIGNMENT)
        if num_cols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        ws.row_dimensions[1].height = 30
        current_row = 2

    header_row = current_row
    if headers:
        ws.row_dimensions[header_row].height = 24
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            apply_cell_style(
                cell,
                fill=HEADER_FILL,
                font=HEADER_FONT,
                alignment=HEADER_ALIGNMENT,
                border=HEADER_BORDER,
            )
        current_row += 1

    for row_idx, row_data in enumerate(data):
        for col_idx, raw_value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = convert_value(raw_value)
            is_alt = row_idx % 2 == 1
            apply_cell_style(
                cell,
                fill=ALT_ROW_FILL if is_alt else None,
                alignment=DATA_ALIGNMENT,
                border=THIN_BORDER,
            )
        current_row += 1

    data_end_row = current_row - 1

    if column_widths:
        for col_idx, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
    elif num_cols > 0:
        for col_idx in range(1, num_cols + 1):
            max_w = 0.0
            for row in ws.iter_rows(
                min_row=1,
                max_row=current_row - 1,
                min_col=col_idx,
                max_col=col_idx,
            ):
                for cell in row:
                    if cell.value is not None:
                        max_w = max(max_w, estimate_width(str(cell.value)))
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(max_w + 4, 8)

    if headers:
        filter_start = header_row
        filter_end = data_end_row
        col_end = get_column_letter(num_cols)
        ws.auto_filter.ref = f"A{filter_start}:{col_end}{filter_end}"

    if headers:
        ws.freeze_panes = f"A{header_row + 1}"
    elif title:
        ws.freeze_panes = "A2"

    for merge_range in merge_cells:
        ws.merge_cells(merge_range)

    return ws


def main():
    if len(sys.argv) < 2:
        print("Usage: python writer.py <input.json> [output.xlsx]", file=sys.stderr)
        print("", file=sys.stderr)
        print("Arguments:", file=sys.stderr)
        print(
            "  input.json   - JSON file describing Excel structure",
            file=sys.stderr,
        )
        print(
            "  output.xlsx  - Optional: output path (default: <input_stem>.xlsx)",
            file=sys.stderr,
        )
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None

    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    try:
        with open(input_path, encoding="utf-8") as f:
            config = json.load(f)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON format: {e}", file=sys.stderr)
        sys.exit(1)

    if not config.get("sheets"):
        print(
            "Error: 'sheets' array is required and must not be empty",
            file=sys.stderr,
        )
        sys.exit(1)

    if output_path is None:
        output_path = input_path.parent / f"{input_path.stem}.xlsx"

    if output_path.suffix.lower() not in (".xlsx", ".xlsm"):
        output_path = output_path.with_suffix(".xlsx")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_config in config["sheets"]:
        if not sheet_config.get("data") and not sheet_config.get("headers"):
            continue
        create_sheet(wb, sheet_config)

    if not wb.sheetnames:
        print("Error: No valid sheet data found", file=sys.stderr)
        sys.exit(1)

    wb.save(output_path)
    print(f"Success: {output_path}")


if __name__ == "__main__":
    main()
