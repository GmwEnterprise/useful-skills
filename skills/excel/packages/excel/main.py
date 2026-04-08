#!/usr/bin/env python3
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.descriptors import sequence as _sequence_module

_original_convert = _sequence_module._convert  # type: ignore[attr-defined]


def _patched_convert(expected_type, value):
    try:
        return _original_convert(expected_type, value)
    except TypeError:
        return None


_sequence_module._convert = _patched_convert  # type: ignore[attr-defined]

import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402


def convert_value(value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.isoformat()

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    return str(value)


def get_cell_type(cell) -> str:
    if cell.value is None:
        return "empty"

    if cell.is_date:
        return "date"

    if isinstance(cell.value, bool):
        return "boolean"

    if isinstance(cell.value, (int, float)):
        return "number"

    return "string"


def get_merged_cell_value(sheet, row: int, col: int, cell) -> tuple[Any, bool]:
    for merged_range in sheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            if row == merged_range.min_row and col == merged_range.min_col:
                return cell.value, False
            else:
                return None, True

    return cell.value, False


def sheet_to_dict(sheet) -> dict:
    max_row = sheet.max_row
    max_col = sheet.max_column

    if max_row == 0 or max_col == 0:
        return {
            "name": sheet.title,
            "data": [],
            "dimensions": {"rows": 0, "columns": 0},
        }

    data = []
    merge_info = []

    for row_idx in range(1, max_row + 1):
        row_data = []
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row_idx, col_idx)
            value, is_merged = get_merged_cell_value(sheet, row_idx, col_idx, cell)

            cell_info = {
                "value": convert_value(value),
                "type": get_cell_type(cell) if not is_merged else "merged",
                "column": get_column_letter(col_idx),
                "row": row_idx,
            }

            row_data.append(cell_info)

        data.append(row_data)

    for merged_range in sheet.merged_cells.ranges:
        merge_info.append(
            {
                "range": str(merged_range),
                "start_row": merged_range.min_row,
                "start_col": merged_range.min_col,
                "end_row": merged_range.max_row,
                "end_col": merged_range.max_col,
            }
        )

    return {
        "name": sheet.title,
        "data": data,
        "dimensions": {"rows": max_row, "columns": max_col},
        "merged_cells": merge_info,
    }


def sheet_to_markdown(sheet) -> str:
    lines = [f"## {sheet.title}", ""]
    max_row = sheet.max_row
    max_col = sheet.max_column

    if max_row == 0 or max_col == 0:
        lines.append("*（空表格）*")
        return "\n".join(lines)

    data = []
    for row_idx in range(1, max_row + 1):
        row_data = []
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row_idx, col_idx)
            value, _ = get_merged_cell_value(sheet, row_idx, col_idx, cell)
            row_data.append(format_cell_value(value))
        data.append(row_data)

    if not data:
        lines.append("*（无数据）*")
        return "\n".join(lines)

    header = data[0]
    lines.append(
        "| " + " | ".join(str(cell) if cell else " " for cell in header) + " |"
    )
    lines.append("| " + " | ".join("---" for _ in header) + " |")

    for row in data[1:]:
        lines.append(
            "| " + " | ".join(str(cell) if cell else " " for cell in row) + " |"
        )

    return "\n".join(lines)


def format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return (
            value.strftime("%Y-%m-%d %H:%M:%S")
            if value.hour or value.minute or value.second
            else value.strftime("%Y-%m-%d")
        )
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def main():
    if len(sys.argv) < 2:
        print("Usage: python main.py <excel_file> [sheet_name]", file=sys.stderr)
        print("\nArguments:", file=sys.stderr)
        print(
            "  excel_file  - Path to the Excel file (.xlsx or .xlsm)",
            file=sys.stderr,
        )
        print(
            "  sheet_name  - Optional: Name of specific sheet to read",
            file=sys.stderr,
        )
        sys.exit(1)

    file_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        path = Path(file_path)
        if not path.exists():
            print(f"Error: Excel file not found: {file_path}", file=sys.stderr)
            sys.exit(1)

        if path.suffix.lower() not in [".xlsx", ".xlsm"]:
            suffix = path.suffix
            msg = f"Error: Unsupported file format: {suffix}. "
            msg += "Only .xlsx and .xlsm are supported."
            print(msg, file=sys.stderr)
            sys.exit(1)

        workbook = openpyxl.load_workbook(path, data_only=True)

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames)
                msg = (
                    f"Error: Sheet '{sheet_name}' not found. "
                    f"Available sheets: {available}"
                )
                print(msg, file=sys.stderr)
                sys.exit(1)
            sheets = [workbook[sheet_name]]
        else:
            sheets = list(workbook)

        markdown_lines = [f"# {path.name}", ""]
        json_data = {"file": str(path.absolute()), "sheets": []}

        for sheet in sheets:
            markdown_lines.append(sheet_to_markdown(sheet))
            markdown_lines.append("")
            json_data["sheets"].append(sheet_to_dict(sheet))

        workbook.close()

        output_dir = path.parent
        base_name = path.stem
        md_file = output_dir / f"{base_name}.excel_reader.md"
        json_file = output_dir / f"{base_name}.excel_reader.json"

        md_file.write_text("\n".join(markdown_lines), encoding="utf-8")
        json_file.write_text(
            json.dumps(json_data, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        print(f"Success: {path.name}")
        print(f"Markdown: {md_file}")
        print(f"JSON: {json_file}")

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
