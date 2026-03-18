#!/usr/bin/env python3
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.descriptors import sequence as _sequence_module

_original_convert = _sequence_module._convert  # type: ignore[attr-defined]


def _patched_convert(expected_type, value):
    try:
        return _original_convert(expected_type, value)
    except TypeError:
        return None


_sequence_module._convert = _patched_convert  # type: ignore[attr-defined]

import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402


def convert_value(value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.isoformat()

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    return str(value)


def get_cell_type(cell) -> str:
    if cell.value is None:
        return "empty"

    if cell.is_date:
        return "date"

    if isinstance(cell.value, bool):
        return "boolean"

    if isinstance(cell.value, (int, float)):
        return "number"

    return "string"


def get_merged_cell_value(sheet, row: int, col: int, cell) -> tuple[Any, bool]:
    for merged_range in sheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            if row == merged_range.min_row and col == merged_range.min_col:
                return cell.value, False
            else:
                return None, True

    return cell.value, False


def sheet_to_dict(sheet) -> dict:
    max_row = sheet.max_row
    max_col = sheet.max_column

    if max_row == 0 or max_col == 0:
        return {
            "name": sheet.title,
            "data": [],
            "dimensions": {"rows": 0, "columns": 0},
        }

    data = []
    merge_info = []

    for row_idx in range(1, max_row + 1):
        row_data = []
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row_idx, col_idx)
            value, is_merged = get_merged_cell_value(sheet, row_idx, col_idx, cell)

            cell_info = {
                "value": convert_value(value),
                "type": get_cell_type(cell) if not is_merged else "merged",
                "column": get_column_letter(col_idx),
                "row": row_idx,
            }

            row_data.append(cell_info)

        data.append(row_data)

    for merged_range in sheet.merged_cells.ranges:
        merge_info.append(
            {
                "range": str(merged_range),
                "start_row": merged_range.min_row,
                "start_col": merged_range.min_col,
                "end_row": merged_range.max_row,
                "end_col": merged_range.max_col,
            }
        )

    return {
        "name": sheet.title,
        "data": data,
        "dimensions": {"rows": max_row, "columns": max_col},
        "merged_cells": merge_info,
    }


def read_excel(file_path: str, sheet_name: str | None = None) -> dict:
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    if path.suffix.lower() not in [".xlsx", ".xlsm"]:
        raise ValueError(
            f"Unsupported file format: {path.suffix}. "
            "Only .xlsx and .xlsm are supported."
        )

    workbook = openpyxl.load_workbook(path, data_only=True)

    result = {"file": str(path.absolute()), "sheets": []}

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {available}"
            )

        sheet = workbook[sheet_name]
        result["sheets"].append(sheet_to_dict(sheet))
    else:
        for sheet in workbook:
            result["sheets"].append(sheet_to_dict(sheet))

    workbook.close()

    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python main.py <excel_file> [sheet_name]", file=sys.stderr)
        print("\nArguments:", file=sys.stderr)
        print(
            "  excel_file  - Path to the Excel file (.xlsx or .xlsm)", file=sys.stderr
        )
        print(
            "  sheet_name  - Optional: Name of specific sheet to read", file=sys.stderr
        )
        sys.exit(1)

    file_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        result = read_excel(file_path, sheet_name)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
